import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Caminho do arquivo de entrada da planilha base
caminho_planilha_base = r"C:\\Users\\MICRO\\OneDrive\\Área de Trabalho\\Programação\\WorkSpace-PYTHON\\WorkSpaceGeral\\Scripts simples prontos\\Script base tallyta\\BASE REFI AGOSTO.xlsx"

# Ler a planilha base
df_base = pd.read_excel(caminho_planilha_base)

# Definir o tamanho do grupo (número de linhas por planilha)
tamanho_grupo = 50

# Definir quantas vezes a planilha será dividida
num_divisoes = 3  # Substitua pelo número desejado

# Adicionar uma coluna "Divisão" à planilha base
df_base["Divisão"] = ""

# Diretório onde as novas planilhas serão criadas
diretorio_saida = r"C:\\Users\\MICRO\\OneDrive\\Área de Trabalho\\Programação\\WorkSpace-PYTHON\\WorkSpaceGeral\\Scripts simples prontos\\Script base tallyta\\Planilhas"

# Lista para armazenar as últimas linhas de cada grupo
ultimas_linhas = []

# Dividir a planilha em grupos e criar planilhas separadas
for j in range(num_divisoes):
    inicio = j * tamanho_grupo
    fim = inicio + tamanho_grupo
    grupo = df_base[inicio:fim]

    # Atualizar a coluna "Divisão" na planilha base
    df_base.loc[inicio:fim, "Divisão"] = f"Divisão {j + 1}"

    # Criar uma nova planilha para o grupo
    nome_planilha = os.path.join(diretorio_saida, f"planilha_divisao_{j + 1}.xlsx")
    grupo.to_excel(nome_planilha, index=False)
    print(f"Planilha {nome_planilha} criada.")

    # Armazenar a última linha do grupo
    ultima_linha_grupo = inicio + tamanho_grupo - 1
    ultimas_linhas.append(ultima_linha_grupo)

# Salvar a planilha base com a coluna "Divisão" atualizada
df_base.to_excel(caminho_planilha_base, index=False)
print("Planilha base atualizada com a coluna 'Divisão'.")


# Estilizar a coluna "Divisão" na planilha base com cor de fundo vermelha e texto branco
def estilizar_planilha_base():
    wb = load_workbook(caminho_planilha_base)
    ws = wb.active

    fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    font = Font(color="FFFFFF")

    for ultima_linha in ultimas_linhas:
        cell = ws.cell(row=ultima_linha + 2, column=ws.max_column)
        cell.fill = fill
        cell.font = font

    wb.save(caminho_planilha_base)


# Estilizar a coluna "Divisão" na planilha base
estilizar_planilha_base()
print("Coluna 'Divisão' estilizada na planilha base.")

print("Processo concluído.")
