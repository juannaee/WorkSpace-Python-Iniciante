import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import tkinter as tk
from tkinter import filedialog


class PlanilhaDividerEstilizerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Script Juan/Base")

        self.label1 = tk.Label(root, text="Caminho da Planilha Base:")
        self.label1.pack()

        self.entry1 = tk.Entry(root)
        self.entry1.pack()

        self.browse_button1 = tk.Button(
            root, text="Procurar", command=self.browse_file1
        )
        self.browse_button1.pack()

        self.label2 = tk.Label(root, text="Número de Consultores:")
        self.label2.pack()

        self.entry2 = tk.Entry(root)
        self.entry2.pack()

        self.label3 = tk.Label(root, text="Caminho do Diretório de Saída:")
        self.label3.pack()

        self.entry3 = tk.Entry(root)
        self.entry3.pack()

        self.browse_button2 = tk.Button(
            root, text="Procurar", command=self.browse_directory
        )
        self.browse_button2.pack()

        self.run_button = tk.Button(root, text="Executar", command=self.run_script)
        self.run_button.pack()

    def browse_file1(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        self.entry1.delete(0, tk.END)
        self.entry1.insert(0, file_path)

    def browse_directory(self):
        directory_path = filedialog.askdirectory()
        self.entry3.delete(0, tk.END)
        self.entry3.insert(0, directory_path)

    def run_script(self):
        caminho_planilha_base = self.entry1.get()
        df_base = pd.read_excel(caminho_planilha_base)

        tamanho_grupo = 50
        num_divisoes = int(self.entry2.get())
        df_base["Divisão"] = ""

        diretorio_saida = self.entry3.get()

        ultimas_linhas = []

        for j in range(num_divisoes):
            inicio = j * tamanho_grupo
            fim = inicio + tamanho_grupo
            grupo = df_base[inicio:fim]
            df_base.loc[inicio:fim, "Divisão"] = f"Divisão {j + 1}"
            nome_planilha = os.path.join(
                diretorio_saida, f"planilha_divisao_{j + 1}.xlsx"
            )
            grupo.to_excel(nome_planilha, index=False)
            ultima_linha_grupo = inicio + tamanho_grupo - 1
            ultimas_linhas.append(ultima_linha_grupo)

        df_base.to_excel(caminho_planilha_base, index=False)

        wb = load_workbook(caminho_planilha_base)
        ws = wb.active
        fill = PatternFill(
            start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
        )
        font = Font(color="FFFFFF")

        for ultima_linha in ultimas_linhas:
            cell = ws.cell(row=ultima_linha + 2, column=ws.max_column)
            cell.fill = fill
            cell.font = font

        wb.save(caminho_planilha_base)

        result_label = tk.Label(self.root, text="Processo concluído.")
        result_label.pack()


if __name__ == "__main__":
    root = tk.Tk()
    app = PlanilhaDividerEstilizerApp(root)
    root.mainloop()
